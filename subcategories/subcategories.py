import openpyxl as op

filename = 'Бланк заказа.xlsx'
subcategories_dict = {}

wb = op.load_workbook(filename, data_only=True)
sheet = wb.active

max_rows = sheet.max_row

for i in range(7, max_rows+1):
    sku = sheet.cell(row=i, column=2).value
    subcategory = sheet.cell(row=i, column=12).value

    if not sku:
        continue

    if subcategory not in subcategories_dict:
        subcategories_dict[subcategory] = [sku]
    else:
        subcategories_dict[subcategory].append(sku)

sorteddict = dict(sorted(subcategories_dict.items()))

print(sorteddict)

with open('subcategories.ini', 'w') as myfile:

    for key, value in sorteddict.items():
        string_values = ', '.join(value)
        string_to_write = key + ' = ' + string_values + '\n'
        myfile.write(string_to_write)